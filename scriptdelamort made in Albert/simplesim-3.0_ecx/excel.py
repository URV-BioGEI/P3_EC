#!/usr/bin/env python
# -*- coding: utf-8 -*-

#from openpyxl import *
from openpyxl import Workbook
import fileinput
import re
import sys
import os

wb = Workbook()
ws = wb.active
#arguments: cache a usar	nombre de modificacions 	parametres de l'execucio[]   
#
#estructura
#
#	tamany1		cache
#	tamany2		cache
#	tamany3		cache
#	tamany4		cache

#primer: tamaño_total en bytes / (tamany de bloc* associativitat) 

tamany = 1
app = 2
cra = 3
two = 4
vor = 5
zip = 6
i = 0
fila_actual=1
argument_actual=3
argument = " "
cadena_busqueda = "{0}.miss_rate".format(sys.argv[1])
#escriure les capcaleres
ws.cell(row=1, column=tamany).value = 'param'
ws.cell(row=1, column=app).value = 'applu'
ws.cell(row=1, column=cra).value = 'crafty'
ws.cell(row=1, column=two).value = 'twolf'
ws.cell(row=1, column=vor).value = 'vortex'
ws.cell(row=1, column=zip).value = 'zip'


#aqui comencarem a escriure la informacio. primer executarem el script que crea el fitxer de sortida per un tamany, despres analitzem el resultat i ho plasmem a la taula, seguidament repetirem el proces per tots els tamanys
while(i < int(sys.argv[2])):	
	t_actual=sys.argv[argument_actual]
	
	#________________________CREACIO DE FITXER_____________________________________
	if sys.argv[1] == "dl1":
		#"sortida[i] tamany:32:4:l 0 0" 
		arguments = "sortida{0} {1} 0 0".format(str(i), str(t_actual))
	elif sys.argv[1] == "il1":
		#"sortida[i] 0 tamany:16:1:l 0" 
		arguments = "sortida{0} 0 {1} 0".format(str(i), str(t_actual))
	elif sys.argv[1] == "ul2":
		#"sortida[i] 0 0 tamany:64:4:l" 
		arguments = "sortida{0} 0 0 {1}".format(str(i), str(t_actual))

	os.system("./Execucio_benchmarks.sh "+arguments)
	
	
	
	
	#________________________LECTURA I ANALISI DE FITXER_____________________________________
	
	for line in fileinput.input("sortida{0}_sim_app".format(str(i))): #applu
	   	
	 	m = re.match(cadena_busqueda, line) # que trobi la expresio dl1.miss_rate, il1.miss_rate  o ul2.miss_rate 
		
		if (m):
			
			ini=12
			while(not line[ini].isdigit()):	#cerquem el primer digit de la linea
				ini=ini+1
			fi_cadena=ini
			while(line[fi_cadena]!='#'):	# cerquem el coixinet que marca el final del valor
				fi_cadena=fi_cadena+1
			
			fi_cadena=fi_cadena-1
			info=line[ini:fi_cadena]	#extraiem la informacio compresa entre el primer digit i la posicio anterior al coixinet					

			c = ws.cell(row = fila_actual, column = tamany) #situem la casella corresponent a la fila actual amb el tamany 
			c.value = t_actual	#escribim el tamany actual
			c = ws.cell(row = fila_actual, column = app)	#situem la casella corresponent a la fila actual amb la cache
			c.value = info		#escribim el miss_rate
			fila_actual = fila_actual + 1
			fileinput.close()
			break
	for line in fileinput.input("sortida{0}_sim_cra".format(str(i))): #crafty
	   	
	 	m = re.match(cadena_busqueda, line) # que trobi la expresio dl1.miss_rate, il1.miss_rate  o ul2.miss_rate 
		
		if (m):
			
			ini=12
			while(not line[ini].isdigit()):	#cerquem el primer digit de la linea
				ini=ini+1
			fi_cadena=ini
			while(line[fi_cadena]!='#'):	# cerquem el coixinet que marca el final del valor
				fi_cadena=fi_cadena+1
			
			fi_cadena=fi_cadena-1
			info=line[ini:fi_cadena]	#extraiem la informacio compresa entre el primer digit i la posicio anterior al coixinet
								

			c = ws.cell(row = fila_actual, column = tamany) #situem la casella corresponent a la fila actual amb el tamany 
			c.value = t_actual	#escribim el tamany actual
			c = ws.cell(row = fila_actual, column = cra)	#situem la casella corresponent a la fila actual amb la cache
			c.value = info		#escribim el miss_rate
			fila_actual = fila_actual + 1
			fileinput.close()
			break
	for line in fileinput.input("sortida{0}_sim_two".format(str(i))): #twolf
	   	
	 	m = re.match(cadena_busqueda, line) # que trobi la expresio dl1.miss_rate, il1.miss_rate  o ul2.miss_rate 
		
		if (m):
			
			ini=12
			while(not line[ini].isdigit()):	#cerquem el primer digit de la linea
				ini=ini+1
			fi_cadena=ini
			while(line[fi_cadena]!='#'):	# cerquem el coixinet que marca el final del valor
				fi_cadena=fi_cadena+1
			
			fi_cadena=fi_cadena-1
			info=line[ini:fi_cadena]	#extraiem la informacio compresa entre el primer digit i la posicio anterior al coixinet
								

			c = ws.cell(row = fila_actual, column = tamany) #situem la casella corresponent a la fila actual amb el tamany 
			c.value = t_actual	#escribim el tamany actual
			c = ws.cell(row = fila_actual, column = two)	#situem la casella corresponent a la fila actual amb la cache
			c.value = info		#escribim el miss_rate
			fila_actual = fila_actual + 1
			fileinput.close()
			break
	for line in fileinput.input("sortida{0}_sim_vor.in".format(str(i))): #vortex
	   	
	 	m = re.match(cadena_busqueda, line) # que trobi la expresio dl1.miss_rate, il1.miss_rate  o ul2.miss_rate 
		
		if (m):
			
			ini=12
			while(not line[ini].isdigit()):	#cerquem el primer digit de la linea
				ini=ini+1
			fi_cadena=ini
			while(line[fi_cadena]!='#'):	# cerquem el coixinet que marca el final del valor
				fi_cadena=fi_cadena+1
			
			fi_cadena=fi_cadena-1
			info=line[ini:fi_cadena]	#extraiem la informacio compresa entre el primer digit i la posicio anterior al coixinet
								

			c = ws.cell(row = fila_actual, column = tamany) #situem la casella corresponent a la fila actual amb el tamany 
			c.value = t_actual	#escribim el tamany actual
			c = ws.cell(row = fila_actual, column = vor)	#situem la casella corresponent a la fila actual amb la cache
			c.value = info		#escribim el miss_rate
			fila_actual = fila_actual + 1
			fileinput.close()
			break
	for line in fileinput.input("sortida{0}_sim_zip.in".format(str(i))): #gzip
	   	
	 	m = re.match(cadena_busqueda, line) # que trobi la expresio dl1.miss_rate, il1.miss_rate  o ul2.miss_rate 
		
		if (m):
			
			ini=12
			while(not line[ini].isdigit()):	#cerquem el primer digit de la linea
				ini=ini+1
			fi_cadena=ini
			while(line[fi_cadena]!='#'):	# cerquem el coixinet que marca el final del valor
				fi_cadena=fi_cadena+1
			
			fi_cadena=fi_cadena-1
			info=line[ini:fi_cadena]	#extraiem la informacio compresa entre el primer digit i la posicio anterior al coixinet
								

			c = ws.cell(row = fila_actual, column = tamany) #situem la casella corresponent a la fila actual amb el tamany 
			c.value = t_actual	#escribim el tamany actual
			c = ws.cell(row = fila_actual, column = zip)	#situem la casella corresponent a la fila actual amb la cache
			c.value = info		#escribim el miss_rate
			fila_actual = fila_actual + 1
			fileinput.close()
			break
	
	i=i+1
	argument_actual=argument_actual+1

	
wb.save('resultats.xlsx')	
	

	


